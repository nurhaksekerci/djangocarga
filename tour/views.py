from gettext import translation
from django.http import HttpResponse, JsonResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, logout
from django.contrib import messages
from django.urls import reverse
from django.utils.translation import gettext_lazy as _
from django.apps import apps
from django.core.cache import cache
from django.db.models import Prefetch
from django.db.models.functions import Coalesce
from django import forms
from django.forms import ModelChoiceField, ModelMultipleChoiceField, ChoiceField
from django.contrib.auth.decorators import login_required
from tour.services import sms
from tour.models import (
    CustomUser, Operation, OperationCustomer, OperationDay, 
    OperationItem, OperationSalesPrice, OperationSubItem, 
    Currency, City, District, Neighborhood, Support, VehicleType, 
    BuyerCompany, Tour, NoVehicleTour, Transfer, Hotel, 
    Museum, Activity, Guide, VehicleSupplier, ActivitySupplier,
    VehicleCost
)

from .services import LoginService, PasswordResetService, sms
from .forms import (
    CurrencyForm, CityForm, DistrictForm, NeighborhoodForm, 
    OperationItemActivityForm, OperationItemNoVehicleGuideForm, OperationItemNoVehicleTourForm, 
    OperationItemVehicleForm, OperationSubItemActivityForm, 
    OperationSubItemGuideForm, OperationSubItemHotelForm, 
    OperationSubItemMuseumForm, OperationSubItemOtherPriceForm, 
    OperationSubItemTourForm, OperationSubItemTransferForm, SendSmsForm, SupportForm,
    VehicleTypeForm, BuyerCompanyForm, TourForm, NoVehicleTourForm, 
    TransferForm, HotelForm, MuseumForm, ActivityForm, GuideForm, 
    VehicleSupplierForm, ActivitySupplierForm, VehicleCostForm, 
    ActivityCostForm, OperationForm, OperationCustomerForm, 
    OperationSalesPriceForm
)

from datetime import datetime, timedelta
from django.utils import timezone
from django.db import transaction
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger


def send_sms(request):
    if request.method == 'POST':
        form = SendSmsForm(request.POST)
        if form.is_valid():
            user = form.cleaned_data['users']
            message = form.cleaned_data['message']
            full_message = f"{user.first_name}, Mesajınız: {message} Gönderen: {request.user.first_name} {request.user.last_name} {request.user.role.title()}"
            sms(user.phone, full_message)
            messages.success(request, 'SMS başarıyla gönderildi.')
            return redirect('tour:send_sms')
    else:
        form = SendSmsForm()
    return render(request, 'header/send_sms.html', {'form': form})


#Authentications
def password_reset_request(request):
    """Parola sıfırlama kodu gönderme sayfası"""
    if request.method == 'POST':
        phone = request.POST.get('phone')
        user = CustomUser.objects.get(phone=phone)
        if not user:
            messages.error(request, _('Bu telefon numarasına sahip bir kullanıcı bulunamadı.'))
            return render(request, 'tour/password-reset.html')
        success, message = PasswordResetService.send_reset_code(phone)
        
        if success:
            messages.success(request, message)
            return redirect('tour:password_reset_verify', phone=phone)
        else:
            messages.error(request, message)
    
    return render(request, 'auth/password-reset.html')

def password_reset_verify(request, phone):
    """Parola sıfırlama kodunu doğrulama sayfası"""
    try:
        user = CustomUser.objects.get(phone=phone)
        if request.method == 'POST':
            code = request.POST.get('code')
            new_password = request.POST.get('password1')
            new_password2 = request.POST.get('password2')
            
            if not code:
                messages.error(request, _('Lütfen doğrulama kodunu girin.'))
                return render(request, 'tour/password_reset_verify.html', {'user': user})
            
            if not new_password or not new_password2:
                messages.error(request, _('Lütfen yeni parolanızı girin.'))
                return render(request, 'tour/password_reset_verify.html', {'user': user})
            
            if new_password != new_password2:
                messages.error(request, _('Parolalar eşleşmiyor.'))
                return render(request, 'tour/password_reset_verify.html', {'user': user})
            
            if len(new_password) < 8:
                messages.error(request, _('Parola en az 8 karakter olmalıdır.'))
                return render(request, 'tour/password_reset_verify.html', {'user': user})
            
            success, message = PasswordResetService.reset_password(user.phone, code, new_password)
            
            if success:
                messages.success(request, message)
                return redirect('tour:login')
            else:
                messages.error(request, message)
        
        return render(request, 'auth/password_reset_verify.html', {'user': user})
    except CustomUser.DoesNotExist:
        messages.error(request, _('Kullanıcı bulunamadı.'))
        return redirect('auth:password_reset_request')

def login_view(request):
    if request.user.is_authenticated:
            return redirect('tour:jobs')
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        next_url = request.POST.get('next')
        user = LoginService.authenticate_user(username, password)
        
        if user is not None:
            if user.is_active:
                login(request, user)
                messages.success(request, _('Başarıyla giriş yaptınız.'))   
                # Kullanıcı rolüne göre yönlendirme
                if next_url:
                    return redirect(next_url)
                elif user.role == 'admin':
                    return redirect('admin:index')
                else:
                    return redirect('tour:jobs')
            else:
                messages.error(request, _('Hesabınız aktif değil. Lütfen yönetici ile iletişime geçin.'))
        else:
            messages.error(request, _('Geçersiz kullanıcı adı veya şifre.'))
    
    return render(request, 'auth/login.html')

def logout_view(request):
    logout(request)
    messages.success(request, _('Başarıyla çıkış yaptınız.'))
    return redirect('tour:login')

#Generic Views
@login_required
def generic_list_view(request, model):
    try:
        model_class = apps.get_model(app_label="tour", model_name=model)
        objects = model_class.objects.all()
        
        # Arama işlemi
        search_query = request.GET.get('search', '')
        if search_query:
            from django.db.models import Q
            search_filters = Q()
            for field in model_class._meta.fields:
                if field.get_internal_type() in ['CharField', 'TextField']:
                    search_filters |= Q(**{f"{field.name}__icontains": search_query})
                elif field.get_internal_type() in ['ForeignKey']:
                    search_filters |= Q(**{f"{field.name}__name__icontains": search_query})
            objects = objects.filter(search_filters)
        
        # Form sınıfını belirle
        form_classes = {
            'Currency': CurrencyForm,
            'VehicleType': VehicleTypeForm,
            'BuyerCompany': BuyerCompanyForm,
            'Tour': TourForm,
            'NoVehicleTour': NoVehicleTourForm,
            'Transfer': TransferForm,
            'Hotel': HotelForm,
            'Museum': MuseumForm,
            'Activity': ActivityForm,
            'Guide': GuideForm,
            'VehicleSupplier': VehicleSupplierForm,
            'ActivitySupplier': ActivitySupplierForm,
            'VehicleCost': VehicleCostForm,
        }
        
        form_class = form_classes.get(model)
        if not form_class:
            raise ValueError(f'Form sınıfı bulunamadı: {model}')
            
        if request.method == 'POST':
            form = form_class(request.POST, request.FILES)
            if form.is_valid():
                form.save()
                messages.success(request, 'Kayıt başarıyla oluşturuldu.')
                
                # HTMX isteği ise sadece form şablonunu döndür
                if request.headers.get('HX-Request'):
                    return render(request, 'generic/generic_form.html', {'form': form_class()})
                    
                return redirect('tour:list', model=model)
        else:
            form = form_class()
            
        # Sayfalama işlemi
        paginator = Paginator(objects, 100)  # Her sayfada 100 kayıt göster
        page = request.GET.get('page')
        
        try:
            objects = paginator.page(page)
        except PageNotAnInteger:
            objects = paginator.page(1)
        except EmptyPage:
            objects = paginator.page(paginator.num_pages)
            
        context = {
            'page_title': f'{model} Listesi',
            'model': model,
            'form': form,
            'objects': objects,
            'fields': model_class._meta.fields,
            'detail_url': 'tour:detail',
            'update_url': 'tour:update',
            'delete_url': 'tour:delete',
            'search_query': search_query,  # Arama sorgusunu context'e ekle
        }
        
        # HTMX isteği ise sadece tablo şablonunu döndür
        if request.headers.get('HX-Request'):
            return render(request, 'generic/generic_table.html', context)
            
        return render(request, 'generic/generic.html', context)
        
    except (LookupError, ValueError) as e:
        messages.error(request, f'Model bulunamadı: {str(e)}')
        return render(request, 'tour/error.html', {'error': 'Model bulunamadı'})


@login_required
def generic_update_view(request, model, pk):
    try:
        model_class = apps.get_model(app_label="tour", model_name=model)
        object = model_class.objects.get(id=pk)
        
        # Form sınıfını belirle
        form_classes = {
            'Currency': CurrencyForm,
            'VehicleType': VehicleTypeForm,
            'BuyerCompany': BuyerCompanyForm,
            'Tour': TourForm,
            'NoVehicleTour': NoVehicleTourForm,
            'Transfer': TransferForm,
            'Hotel': HotelForm,
            'Museum': MuseumForm,
            'Activity': ActivityForm,
            'Guide': GuideForm,
            'VehicleSupplier': VehicleSupplierForm,
            'ActivitySupplier': ActivitySupplierForm,
            'VehicleCost': VehicleCostForm,
        }
        
        form_class = form_classes.get(model)
        if not form_class:
            raise ValueError(f'Form sınıfı bulunamadı: {model}')
            
        if request.method == 'POST':
            form = form_class(request.POST, request.FILES, instance=object)
            if form.is_valid():
                form.save()
                messages.success(request, 'Kayıt başarıyla güncellendi.')
                
                # HTMX isteği ise hem tablo hem de form verilerini döndür
                if request.headers.get('HX-Request'):
                    # Tablo verilerini hazırla
                    table_context = {
                        'objects': model_class.objects.all(),
                        'fields': model_class._meta.fields,
                        'detail_url': 'tour:detail',
                        'update_url': 'tour:update',
                        'delete_url': 'tour:delete',
                        'model': model
                    }
                    
                    # Tablo HTML'ini oluştur
                    table_html = render(request, 'generic/generic_table.html', table_context).content.decode('utf-8')
                    
                    # Form verilerini hazırla
                    form_context = {
                        'form': form_class(),
                        'model': model
                    }
                    
                    # Form HTML'ini oluştur
                    form_html = render(request, 'generic/generic_form.html', form_context).content.decode('utf-8')
                    
                    # Her iki HTML'i birleştir
                    return HttpResponse(form_html + table_html)
                    
                return redirect('tour:list', model=model)
        else:
            form = form_class(instance=object)  
            
        context = {
            'item': object,
            'page_title': f'{model} Düzenle',
            'model': model,
            'form': form,
            'object': object
        }
        
        # HTMX isteği ise sadece güncelleme form şablonunu döndür
        if request.headers.get('HX-Request'):
            return render(request, 'generic/generic_update_form.html', context)
            
        return render(request, 'generic/generic_update_form.html', context)
        
    except (LookupError, ValueError) as e:
        messages.error(request, f'Model bulunamadı: {str(e)}')
        return render(request, 'tour/error.html', {'error': 'Model bulunamadı'})


@login_required
def generic_delete_view(request, model, pk):
    try:
        # Model sınıfını al
        model_class = apps.get_model(app_label="tour", model_name=model)
        # Nesneyi bul
        obj = get_object_or_404(model_class, pk=pk)
        
        # Nesneyi sil
        obj.delete()
        messages.success(request, 'Kayıt başarıyla silindi.')
        
        # HTMX isteği ise boş yanıt döndür (satır zaten silinecek)
        if request.headers.get('HX-Request'):
            return HttpResponse('')
        
        # Liste sayfasına yönlendir
        return redirect('tour:list', model=model)
        
    except (LookupError, ValueError) as e:
        messages.error(request, f'Model bulunamadı: {str(e)}')
        return render(request, 'tour/error.html', {'error': 'Model bulunamadı'})
    except Exception as e:
        messages.error(request, f'Bir hata oluştu: {str(e)}')
        return render(request, 'tour/error.html', {'error': 'Bir hata oluştu'})


@login_required
def generic_export_view(request, model):
    try:
        model_class = apps.get_model(app_label="tour", model_name=model)
        objects = model_class.objects.all()
        
        # Excel dosyası oluştur
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import HttpResponse
        
        # Yeni bir Excel çalışma kitabı oluştur
        wb = Workbook()
        ws = wb.active
        ws.title = model
        
        # Başlık satırı için stil tanımla
        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Başlık satırını ekle
        headers = [field.verbose_name for field in model_class._meta.fields]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Veri satırlarını ekle
        for row, obj in enumerate(objects, 2):
            for col, field in enumerate(model_class._meta.fields, 1):
                value = getattr(obj, field.name)
                
                # İlişkili alanları işle
                if hasattr(value, '_meta'):  # Eğer değer bir model nesnesi ise
                    value = str(value)
                elif field.name == 'image' and value:
                    value = value.url
                elif field.name == 'date' and value:
                    value = value.strftime('%d.%m.%Y')
                elif field.name == 'time' and value:
                    value = value.strftime('%H:%M')
                elif field.name == 'is_active' and value is not None:
                    value = 'Aktif' if value else 'Pasif'
                elif field.name == 'created_at' and value:
                    value = value.strftime('%d.%m.%Y %H:%M')
                elif field.name == 'updated_at' and value:
                    value = value.strftime('%d.%m.%Y %H:%M')
                
                ws.cell(row=row, column=col, value=value)
        
        # Sütun genişliklerini otomatik ayarla
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Excel dosyasını HTTP yanıtı olarak döndür
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{model.lower()}_export.xlsx"'
        wb.save(response)
        
        return response
        
    except Exception as e:
        messages.error(request, f'Dışa aktarma hatası: {str(e)}')
        return redirect('tour:list', model=model)

@login_required
def generic_detail_view(request, model, pk):
    try:
        # Model sınıfını al
        model_class = apps.get_model(app_label="tour", model_name=model)
        # Nesneyi bul
        obj = get_object_or_404(model_class, pk=pk)
        
        # Form sınıfını belirle
        form_classes = {
            'Currency': CurrencyForm,
            'VehicleType': VehicleTypeForm,
            'BuyerCompany': BuyerCompanyForm,
            'Tour': TourForm,
            'NoVehicleTour': NoVehicleTourForm,
            'Transfer': TransferForm,
            'Hotel': HotelForm,
            'Museum': MuseumForm,
            'Activity': ActivityForm,
            'Guide': GuideForm,
            'VehicleSupplier': VehicleSupplierForm,
            'ActivitySupplier': ActivitySupplierForm,
            'VehicleCost': VehicleCostForm,
            'ActivityCost': ActivityCostForm,
        }
        
        form_class = form_classes.get(model)
        if not form_class:
            raise ValueError(f'Form sınıfı bulunamadı: {model}')
            
        # Form oluştur (salt okunur)
        form = form_class(instance=obj)
        for field in form.fields.values():
            field.disabled = True
            
        context = {
            'item': obj,
            'page_title': f'{model} Detayı',
            'model': model,
            'form': form,
            'fields': model_class._meta.fields,
        }
        
        # HTMX isteği ise sadece form şablonunu döndür
        if request.headers.get('HX-Request'):
            return render(request, 'generic/generic_form.html', context)
            
        return render(request, 'generic/generic_detail.html', context)
        
    except (LookupError, ValueError) as e:
        messages.error(request, f'Model bulunamadı: {str(e)}')
        return render(request, 'tour/error.html', {'error': 'Model bulunamadı'})

@login_required
def generic_create_view(request, model):
    try:
        model_class = apps.get_model(app_label="tour", model_name=model)
        
        # Form sınıfını belirle
        form_classes = {
            'Currency': CurrencyForm,
            'VehicleType': VehicleTypeForm,
            'BuyerCompany': BuyerCompanyForm,
            'Tour': TourForm,
            'NoVehicleTour': NoVehicleTourForm,
            'Transfer': TransferForm,
            'Hotel': HotelForm,
            'Museum': MuseumForm,
            'Activity': ActivityForm,
            'Guide': GuideForm,
            'VehicleSupplier': VehicleSupplierForm,
            'ActivitySupplier': ActivitySupplierForm,
            'VehicleCost': VehicleCostForm,
        }
        
        form_class = form_classes.get(model)
        if not form_class:
            raise ValueError(f'Form sınıfı bulunamadı: {model}')
            
        if request.method == 'POST':
            form = form_class(request.POST, request.FILES)
            if form.is_valid():
                form.save()
                messages.success(request, 'Kayıt başarıyla oluşturuldu.')
                
                # HTMX isteği ise sadece form şablonunu döndür
                if request.headers.get('HX-Request'):
                    return render(request, 'generic/generic_form.html', {'form': form_class()})
                    
                return redirect('tour:list', model=model)
        else:
            form = form_class()
            
        context = {
            'page_title': f'{model} Oluştur',
            'model': model,
            'form': form
        }
        
        # HTMX isteği ise sadece form şablonunu döndür
        if request.headers.get('HX-Request'):
            return render(request, 'generic/generic_form.html', context)
            
        return render(request, 'generic/generic_form.html', context)
        
    except (LookupError, ValueError) as e:
        messages.error(request, f'Model bulunamadı: {str(e)}')
        return render(request, 'tour/error.html', {'error': 'Model bulunamadı'})


#Operasyon İşlemleri
#Operasyon Durumunu Değiştir
@login_required
def toggle_operation(request, operation_id):
    operation = get_object_or_404(Operation, id=operation_id)
    new_status = not operation.is_active
    next_url = request.GET.get('next')
    # Tek bir transaction içinde tüm işlemleri gerçekleştir
    with transaction.atomic():
        # Operasyon durumunu güncelle
        operation.is_active = new_status
        operation.save()

        OperationCustomer.objects.filter(operation=operation).update(is_active=new_status)
        OperationSalesPrice.objects.filter(operation=operation).update(is_active=new_status)
        # İlgili tüm günleri güncelle
        OperationDay.objects.filter(operation=operation).update(is_active=new_status)
        # İlgili tüm öğeleri güncelle
        OperationItem.objects.filter(operation_day__operation=operation).update(is_active=new_status)
        # İlgili tüm alt öğeleri güncelle
        OperationSubItem.objects.filter(operation_item__operation_day__operation=operation).update(is_active=new_status)
    return redirect('tour:operation_list')

#Operasyon Müşteri Durumunu Değiştir
@login_required
def toggle_operation_customer(request, operation_customer_id):
    customer = get_object_or_404(OperationCustomer, id=operation_customer_id)
    customer.is_active = not customer.is_active
    customer.save()
    next_url = request.GET.get('next')
    if next_url:
        return redirect(next_url)
    return redirect('tour:operation_day_create', operation_id=customer.operation.id)

#Operasyon Satış Fiyatı Durumunu Değiştir
@login_required
def toggle_operation_sales_price(request, operation_sales_price_id):
    sales_price = get_object_or_404(OperationSalesPrice, id=operation_sales_price_id)
    sales_price.is_active = not sales_price.is_active
    sales_price.save()
    next_url = request.GET.get('next')
    if next_url:
        return redirect(next_url)
    return redirect('tour:operation_day_create', operation_id=sales_price.operation.id)

#Operasyon Gün Durumunu Değiştir
@login_required
def toggle_operation_day(request, operation_day_id):
    operation_day = get_object_or_404(OperationDay, id=operation_day_id)
    new_status = not operation_day.is_active
    next_url = request.GET.get('next')
    # Tek bir transaction içinde tüm işlemleri gerçekleştir
    with transaction.atomic():
        # Gün durumunu güncelle
        operation_day.is_active = new_status
        operation_day.save()
        
        # İlgili tüm öğeleri güncelle
        OperationItem.objects.filter(operation_day=operation_day).update(is_active=new_status)
        
        # İlgili tüm alt öğeleri güncelle
        OperationSubItem.objects.filter(operation_item__operation_day=operation_day).update(is_active=new_status)
    if next_url:
        return redirect(next_url)
    return redirect('tour:operation_day_create', operation_id=operation_day.operation.id)

#Operasyon Öğe Durumunu Değiştir
@login_required
def toggle_operation_item(request, operation_item_id):
    operation_item = get_object_or_404(OperationItem, id=operation_item_id)
    new_status = not operation_item.is_active
    next_url = request.GET.get('next')

    with transaction.atomic():
        # Öğe durumunu güncelle
        operation_item.is_active = new_status
        operation_item.save()
        
        # İlgili tüm alt öğeleri güncelle
        OperationSubItem.objects.filter(operation_item=operation_item).update(is_active=new_status)
    if next_url:
        return redirect(next_url)
    return redirect('tour:operation_day_create', operation_id=operation_item.operation_day.operation.id)

#Operasyon Alt Öğe Durumunu Değiştir
@login_required
def toggle_operation_sub_item(request, operation_sub_item_id):
    operation_sub_item = get_object_or_404(OperationSubItem, id=operation_sub_item_id)
    operation_sub_item.is_active = not operation_sub_item.is_active
    operation_sub_item.save()
    next_url = request.GET.get('next')
    if next_url:
        return redirect(next_url)
    return redirect('tour:operation_day_create', operation_id=operation_sub_item.operation_item.operation_day.operation.id)

#Operasyon İşlemleri
#Operasyon Görüntüle
def operation(request, operation_id):
    operation = get_object_or_404(Operation, id=operation_id)
    customers = OperationCustomer.objects.filter(operation=operation).select_related('operation')
    sales_prices = OperationSalesPrice.objects.filter(operation=operation).select_related('operation')
    days = OperationDay.objects.filter(operation=operation).select_related('operation')
    items = OperationItem.objects.filter(operation_day__in=days).select_related(
        'operation_day',
        'vehicle_type',
        'vehicle_supplier',
        'no_vehicle_tour',
        'no_vehicle_activity',
        'activity_supplier'
    )
    sub_items = OperationSubItem.objects.filter(operation_item__in=items).select_related(
        'operation_item',
        'tour',
        'transfer',
        'hotel',
        'guide',
        'activity',
        'activity_supplier'
    ).prefetch_related('museums')
    return render(request, 'operation/operation.html', {
        'operation': operation, 
        'customers': customers, 
        'sales_prices': sales_prices, 
        'days': days, 
        'items': items, 
        'sub_items': sub_items
    })

def operation_update(request, operation_id):
    operation = get_object_or_404(Operation, id=operation_id)
    customers = OperationCustomer.objects.filter(operation=operation).select_related('operation')
    sales_prices = OperationSalesPrice.objects.filter(operation=operation).select_related('operation')
    days = OperationDay.objects.filter(operation=operation).select_related('operation')
    items = OperationItem.objects.filter(operation_day__in=days).select_related(
        'operation_day',
        'vehicle_type',
        'vehicle_supplier',
        'no_vehicle_tour',
        'no_vehicle_activity',
        'activity_supplier'
    )
    sub_items = OperationSubItem.objects.filter(operation_item__in=items).select_related(
        'operation_item',
        'tour',
        'transfer',
        'hotel',
        'guide',
        'activity',
        'activity_supplier'
    ).prefetch_related('museums')
    form = OperationForm(instance=operation)
    if request.method == 'POST':
        form = OperationForm(request.POST, instance=operation)
        if form.is_valid():
            form.save()
            return render(request, 'operation/includes/operation_detail.html', {
                'page_title': 'Operasyon Düzenle',
                'form': form,
                'operation': operation,
                'customers': customers,
                'sales_prices': sales_prices,
                'days': days,
                'items': items,
                'sub_items': sub_items
            })
    return render(request, 'operation/forms/form.html', {
        'operation': operation,
        'page_title': 'Operasyon Düzenle',
        'form': form,
        'page_title': 'Operasyon Düzenle',
        'post_url': reverse('tour:operation_update', args=[operation.id])
    })

def operation_customer_update(request, operation_customer_id):
    customer = get_object_or_404(OperationCustomer, id=operation_customer_id)
    operation = get_object_or_404(Operation, id=customer.operation.id)
    customers = OperationCustomer.objects.filter(operation=operation).select_related('operation')
    sales_prices = OperationSalesPrice.objects.filter(operation=operation).select_related('operation')
    days = OperationDay.objects.filter(operation=operation).select_related('operation')
    items = OperationItem.objects.filter(operation_day__in=days).select_related(
        'operation_day',
        'vehicle_type',
        'vehicle_supplier',
        'no_vehicle_tour',
        'no_vehicle_activity',
        'activity_supplier'
    )
    sub_items = OperationSubItem.objects.filter(operation_item__in=items).select_related(
        'operation_item',
        'tour',
        'transfer',
        'hotel',
        'guide',
        'activity',
        'activity_supplier'
    ).prefetch_related('museums')
    form = OperationCustomerForm(instance=customer)
    if request.method == 'POST':
        form = OperationCustomerForm(request.POST, instance=customer)
        if form.is_valid():
            form.save()
            return render(request, 'operation/includes/operation_detail.html', {
                'form': form,
                'customer': customer,
                'operation': operation,
                'customers': customers,
                'sales_prices': sales_prices,
                'days': days,
                'items': items,
                'sub_items': sub_items
            })
        else:
            print(form.errors)
    return render(request, 'operation/forms/form.html', {
        'form': form,
        'customer': customer,
        'operation': operation,
        'page_title': 'Müşteri Düzenle',
        'post_url': reverse('tour:operation_customer_update', args=[customer.id])
    })


def operation_sales_price_update(request, operation_sales_price_id):
    sales_price = get_object_or_404(OperationSalesPrice, id=operation_sales_price_id)
    operation = get_object_or_404(Operation, id=sales_price.operation.id)
    customers = OperationCustomer.objects.filter(operation=operation).select_related('operation')
    sales_prices = OperationSalesPrice.objects.filter(operation=operation).select_related('operation')
    days = OperationDay.objects.filter(operation=operation).select_related('operation')
    items = OperationItem.objects.filter(operation_day__in=days).select_related(
        'operation_day',
        'vehicle_type',
        'vehicle_supplier',
        'no_vehicle_tour',
        'no_vehicle_activity',
        'activity_supplier'
    )
    sub_items = OperationSubItem.objects.filter(operation_item__in=items).select_related(
        'operation_item',
        'tour',
        'transfer',
        'hotel',
        'guide',
        'activity',
        'activity_supplier'
    ).prefetch_related('museums')
    form = OperationSalesPriceForm(instance=sales_price)
    if request.method == 'POST':
        form = OperationSalesPriceForm(request.POST, instance=sales_price)
        if form.is_valid():
            form.save()
            return render(request, 'operation/includes/operation_detail.html', {
                'form': form,
                'sales_price': sales_price,
                'operation': operation,
                'customers': customers,
                'sales_prices': sales_prices,
                'days': days,
                'items': items,
                'sub_items': sub_items
            })
    return render(request, 'operation/forms/form.html', {
        'form': form,
        'sales_price': sales_price,
        'operation': operation,
        'page_title': 'Satış Fiyatı Düzenle',
        'post_url': reverse('tour:operation_sales_price_update', args=[sales_price.id])
    })


def vehicle_item_create(request, operation_day_id):
    day = get_object_or_404(OperationDay, id=operation_day_id)
    operation = get_object_or_404(Operation, id=day.operation.id)
    customers = OperationCustomer.objects.filter(operation=operation).select_related('operation')
    sales_prices = OperationSalesPrice.objects.filter(operation=operation).select_related('operation')
    days = OperationDay.objects.filter(operation=operation).select_related('operation')
    items = OperationItem.objects.filter(operation_day__in=days).select_related(
        'operation_day',
        'vehicle_type',
        'vehicle_supplier',
        'no_vehicle_tour',
        'no_vehicle_activity',
        'activity_supplier'
    )
    sub_items = OperationSubItem.objects.filter(operation_item__in=items).select_related(
        'operation_item',
        'tour',
        'transfer',
        'hotel',
        'guide',
        'activity',
        'activity_supplier'
    ).prefetch_related('museums')
    form = OperationItemVehicleForm()
    if request.method == 'POST':
        form = OperationItemVehicleForm(request.POST)
        if form.is_valid():
            item = form.save(commit=False)
            item.operation_day = day
            item.item_type = "VEHICLE"
            item.save()
            return render(request, 'operation/includes/operation_detail.html', {
                'form': form,
                'operation': operation,
                'customers': customers,
                'sales_prices': sales_prices,
                'days': days,
                'items': items,
                'sub_items': sub_items
            })
        else:
            print(form.errors)
    return render(request, 'operation/forms/form.html', {
        'form': form,
        'day': day,
        'operation': operation,
        'page_title': 'Araç Ekle',
        'post_url': reverse('tour:vehicle_item_create', args=[day.id])
    })

def vehicle_item_update(request, operation_item_id):
    item = get_object_or_404(OperationItem, id=operation_item_id)
    day = get_object_or_404(OperationDay, id=item.operation_day.id)
    operation = get_object_or_404(Operation, id=day.operation.id)
    customers = OperationCustomer.objects.filter(operation=operation).select_related('operation')
    sales_prices = OperationSalesPrice.objects.filter(operation=operation).select_related('operation')
    days = OperationDay.objects.filter(operation=operation).select_related('operation')
    items = OperationItem.objects.filter(operation_day__in=days).select_related(
        'operation_day',
        'vehicle_type',
        'vehicle_supplier',
        'no_vehicle_tour',
        'no_vehicle_activity',
        'activity_supplier'
    )
    sub_items = OperationSubItem.objects.filter(operation_item__in=items).select_related(
        'operation_item',
        'tour',
        'transfer',
        'hotel',
        'guide',
        'activity',
        'activity_supplier'
    ).prefetch_related('museums')
    form = OperationItemVehicleForm(instance=item)
    if request.method == 'POST':
        form = OperationItemVehicleForm(request.POST, instance=item)
        if form.is_valid():
            form.save()
            return render(request, 'operation/includes/operation_detail.html', {
                'form': form,
                'operation': operation,
                'customers': customers,
                'sales_prices': sales_prices,
                'days': days,
                'items': items,
                'sub_items': sub_items
            })
    return render(request, 'operation/forms/form.html', {
        'form': form,
        'day': day,
        'operation': operation,
        'page_title': 'Araç Düzenle',
        'post_url': reverse('tour:vehicle_item_update', args=[item.id])
    })
            
def no_vehicle_activity_item_create(request, operation_day_id):
    day = get_object_or_404(OperationDay, id=operation_day_id)
    operation = get_object_or_404(Operation, id=day.operation.id)
    customers = OperationCustomer.objects.filter(operation=operation).select_related('operation')
    sales_prices = OperationSalesPrice.objects.filter(operation=operation).select_related('operation')
    days = OperationDay.objects.filter(operation=operation).select_related('operation')
    items = OperationItem.objects.filter(operation_day__in=days).select_related(
        'operation_day',
        'vehicle_type',
        'vehicle_supplier',
        'no_vehicle_tour',
        'no_vehicle_activity',
        'activity_supplier'
    )
    sub_items = OperationSubItem.objects.filter(operation_item__in=items).select_related(
        'operation_item',
        'tour',
        'transfer',
        'hotel',
        'guide',
        'activity',
        'activity_supplier'
    ).prefetch_related('museums')
    form = OperationItemActivityForm()
    if request.method == 'POST':    
        form = OperationItemActivityForm(request.POST)
        if form.is_valid():
            item = form.save(commit=False)
            item.operation_day = day
            item.item_type = "NO_VEHICLE_ACTIVITY"
            item.save()
            return render(request, 'operation/includes/operation_detail.html', {
                'form': form,
                'operation': operation,
                'customers': customers,
                'sales_prices': sales_prices,
                'days': days,
                'items': items,
                'sub_items': sub_items
            })
    return render(request, 'operation/forms/form.html', {
        'form': form,
        'day': day,
        'operation': operation,
        'page_title': 'Araçsız Aktivite Ekle',
        'post_url': reverse('tour:no_vehicle_activity_item_create', args=[day.id])
    })

def no_vehicle_activity_item_update(request, operation_item_id):
    item = get_object_or_404(OperationItem, id=operation_item_id)
    day = get_object_or_404(OperationDay, id=item.operation_day.id)
    operation = get_object_or_404(Operation, id=day.operation.id)
    customers = OperationCustomer.objects.filter(operation=operation).select_related('operation')
    sales_prices = OperationSalesPrice.objects.filter(operation=operation).select_related('operation')
    days = OperationDay.objects.filter(operation=operation).select_related('operation')
    items = OperationItem.objects.filter(operation_day__in=days).select_related(
        'operation_day',
        'vehicle_type',
        'vehicle_supplier',
        'no_vehicle_tour',
        'no_vehicle_activity',
        'activity_supplier'
    )
    sub_items = OperationSubItem.objects.filter(operation_item__in=items).select_related(
        'operation_item',
        'tour',
        'transfer',
        'hotel',
        'guide',
        'activity',
        'activity_supplier'
    ).prefetch_related('museums')
    form = OperationItemActivityForm(instance=item)
    if request.method == 'POST':
        form = OperationItemActivityForm(request.POST, instance=item)
        if form.is_valid():
            form.save()
            return render(request, 'operation/includes/operation_detail.html', {
                'form': form,
                'operation': operation,
                'customers': customers,
                'sales_prices': sales_prices,
                'days': days,
                'items': items,
                'sub_items': sub_items
            })
    return render(request, 'operation/forms/form.html', {
        'form': form,
        'day': day,
        'operation': operation,
        'page_title': 'Araçsız Aktivite Düzenle',
        'post_url': reverse('tour:no_vehicle_activity_item_update', args=[item.id])
    })
    
        
def no_vehicle_tour_item_create(request, operation_day_id):
    day = get_object_or_404(OperationDay, id=operation_day_id)
    operation = get_object_or_404(Operation, id=day.operation.id)
    customers = OperationCustomer.objects.filter(operation=operation).select_related('operation')
    sales_prices = OperationSalesPrice.objects.filter(operation=operation).select_related('operation')
    days = OperationDay.objects.filter(operation=operation).select_related('operation')
    items = OperationItem.objects.filter(operation_day__in=days).select_related(
        'operation_day',
        'vehicle_type',
        'vehicle_supplier',
        'no_vehicle_tour',
        'no_vehicle_activity',
        'activity_supplier'
    )
    sub_items = OperationSubItem.objects.filter(operation_item__in=items).select_related(
        'operation_item',
        'tour',
        'transfer',
        'hotel',
        'guide',
        'activity',
        'activity_supplier'
    ).prefetch_related('museums')
    form = OperationItemNoVehicleTourForm()
    if request.method == 'POST':
        form = OperationItemNoVehicleTourForm(request.POST)
        if form.is_valid():
            item = form.save(commit=False)
            item.operation_day = day
            item.item_type = "NO_VEHICLE_TOUR"
            item.save()
            return render(request, 'operation/includes/operation_detail.html', {
                'form': form,
                'operation': operation,
                'customers': customers,
                'sales_prices': sales_prices,
                'days': days,
                'items': items,
                'sub_items': sub_items
            })
    return render(request, 'operation/forms/form.html', {
        'form': form,
        'day': day,
        'operation': operation,
        'page_title': 'Araçsız Tur Ekle',
        'post_url': reverse('tour:no_vehicle_tour_item_create', args=[day.id])
    })

def no_vehicle_tour_item_update(request, operation_item_id):
    item = get_object_or_404(OperationItem, id=operation_item_id)
    day = get_object_or_404(OperationDay, id=item.operation_day.id)
    operation = get_object_or_404(Operation, id=day.operation.id)
    customers = OperationCustomer.objects.filter(operation=operation).select_related('operation')
    sales_prices = OperationSalesPrice.objects.filter(operation=operation).select_related('operation')
    days = OperationDay.objects.filter(operation=operation).select_related('operation')
    items = OperationItem.objects.filter(operation_day__in=days).select_related(
        'operation_day',
        'vehicle_type',
        'vehicle_supplier',
        'no_vehicle_tour',
        'no_vehicle_activity',
        'activity_supplier'
    )
    sub_items = OperationSubItem.objects.filter(operation_item__in=items).select_related(
        'operation_item',
        'tour',
        'transfer',
        'hotel',
        'guide',
        'activity',
        'activity_supplier'
    ).prefetch_related('museums')
    form = OperationItemNoVehicleTourForm(instance=item)
    if request.method == 'POST':
        form = OperationItemNoVehicleTourForm(request.POST, instance=item)
        if form.is_valid():
            form.save()
            return render(request, 'operation/includes/operation_detail.html', {
                'form': form,
                'operation': operation,
                'customers': customers,
                'sales_prices': sales_prices,
                'days': days,
                'items': items,
                'sub_items': sub_items
            })
    return render(request, 'operation/forms/form.html', {
        'form': form,
        'day': day,
        'operation': operation,
        'page_title': 'Araçsız Tur Düzenle',
        'post_url': reverse('tour:no_vehicle_tour_item_update', args=[item.id])
    })

def no_vehicle_guide_item_create(request, operation_day_id):
    day = get_object_or_404(OperationDay, id=operation_day_id)
    operation = get_object_or_404(Operation, id=day.operation.id)
    customers = OperationCustomer.objects.filter(operation=operation).select_related('operation')
    sales_prices = OperationSalesPrice.objects.filter(operation=operation).select_related('operation')
    days = OperationDay.objects.filter(operation=operation).select_related('operation') 
    items = OperationItem.objects.filter(operation_day__in=days).select_related(
        'operation_day',
        'vehicle_type',
        'vehicle_supplier',
        'no_vehicle_tour',
        'no_vehicle_activity',
        'activity_supplier'
    )
    sub_items = OperationSubItem.objects.filter(operation_item__in=items).select_related(
        'operation_item',
        'tour',
        'transfer',
        'hotel',
        'guide',
        'activity',
        'activity_supplier'
    ).prefetch_related('museums')
    form = OperationItemNoVehicleGuideForm()
    if request.method == 'POST':
        form = OperationItemNoVehicleGuideForm(request.POST)
        if form.is_valid():
            item = form.save(commit=False)
            item.operation_day = day
            item.item_type = "NO_VEHICLE_GUIDE"
            item.save()
            return render(request, 'operation/includes/operation_detail.html', {
                'form': form,
                'operation': operation,
                'customers': customers,
                'sales_prices': sales_prices,
                'days': days,
                'items': items,
                'sub_items': sub_items
            })
    return render(request, 'operation/forms/form.html', {
        'form': form,
        'day': day,
        'operation': operation,
        'page_title': 'Araçsız Rehber Ekle',
        'post_url': reverse('tour:no_vehicle_guide_item_create', args=[day.id])
    })

def no_vehicle_guide_item_update(request, operation_item_id):
    item = get_object_or_404(OperationItem, id=operation_item_id)
    day = get_object_or_404(OperationDay, id=item.operation_day.id)
    operation = get_object_or_404(Operation, id=day.operation.id)
    customers = OperationCustomer.objects.filter(operation=operation).select_related('operation')
    sales_prices = OperationSalesPrice.objects.filter(operation=operation).select_related('operation')
    days = OperationDay.objects.filter(operation=operation).select_related('operation')
    items = OperationItem.objects.filter(operation_day__in=days).select_related(
        'operation_day',
        'vehicle_type',
        'vehicle_supplier',
        'no_vehicle_tour',
        'no_vehicle_activity',
        'activity_supplier'
    )
    sub_items = OperationSubItem.objects.filter(operation_item__in=items).select_related(
        'operation_item',
        'tour',
        'transfer',
        'hotel',
        'guide',
        'activity',
        'activity_supplier'
    ).prefetch_related('museums')
    form = OperationItemNoVehicleGuideForm(instance=item)
    if request.method == 'POST':
        form = OperationItemNoVehicleGuideForm(request.POST, instance=item)
        if form.is_valid():
            form.save()
            return render(request, 'operation/includes/operation_detail.html', {
                'form': form,
                'operation': operation,
                'customers': customers,
                'sales_prices': sales_prices,
                'days': days,
                'items': items,
                'sub_items': sub_items
            })
    return render(request, 'operation/forms/form.html', {
        'form': form,
        'operation': operation,
        'page_title': 'Araçsız Rehber Düzenle',
        'post_url': reverse('tour:no_vehicle_guide_item_update', args=[item.id])
    })


def sub_item_tour_create(request, operation_item_id):
    item = get_object_or_404(OperationItem, id=operation_item_id)
    operation = get_object_or_404(Operation, id=item.operation_day.operation.id)
    customers = OperationCustomer.objects.filter(operation=operation).select_related('operation')
    sales_prices = OperationSalesPrice.objects.filter(operation=operation).select_related('operation')
    days = OperationDay.objects.filter(operation=operation).select_related('operation')
    items = OperationItem.objects.filter(operation_day__in=days).select_related(
        'operation_day',
        'vehicle_type',
        'vehicle_supplier',
        'no_vehicle_tour',
        'no_vehicle_activity',
        'activity_supplier'
    )
    sub_items = OperationSubItem.objects.filter(operation_item__in=items).select_related(
        'operation_item',
        'tour',
        'transfer',
        'hotel',
        'guide',
        'activity',
        'activity_supplier'
    ).prefetch_related('museums')
    form = OperationSubItemTourForm()
    if request.method == 'POST':
        form = OperationSubItemTourForm(request.POST)
        if form.is_valid():
            tour = form.save(commit=False)
            tour.operation_item = item
            tour.subitem_type = "TOUR"
            tour.save()
            return render(request, 'operation/includes/operation_detail.html', {
                'form': form,
                'operation': operation,
                'customers': customers,
                'sales_prices': sales_prices,
                'days': days,
                'items': items,
                'sub_items': sub_items
            })
    return render(request, 'operation/forms/form.html', {
        'form': form,
        'operation': operation,
        'page_title': 'Tur Ekle',
        'post_url': reverse('tour:sub_item_tour_create', args=[item.id])
    })

def sub_item_tour_update(request, operation_sub_item_id):
    sub_item = get_object_or_404(OperationSubItem, id=operation_sub_item_id)
    operation = get_object_or_404(Operation, id=sub_item.operation_item.operation_day.operation.id)
    customers = OperationCustomer.objects.filter(operation=operation).select_related('operation')
    sales_prices = OperationSalesPrice.objects.filter(operation=operation).select_related('operation')
    days = OperationDay.objects.filter(operation=operation).select_related('operation')
    items = OperationItem.objects.filter(operation_day__in=days).select_related(
        'operation_day',
        'vehicle_type',
        'vehicle_supplier',
        'no_vehicle_tour',
        'no_vehicle_activity',
        'activity_supplier'
    )
    sub_items = OperationSubItem.objects.filter(operation_item__in=items).select_related(
        'operation_item',
        'tour',
        'transfer',
        'hotel',
        'guide',
        'activity',
        'activity_supplier'
    ).prefetch_related('museums')
    form = OperationSubItemTourForm(instance=sub_item)
    if request.method == 'POST':
        form = OperationSubItemTourForm(request.POST, instance=sub_item)
        if form.is_valid():
            form.save()
            return render(request, 'operation/includes/operation_detail.html', {
                'form': form,
                'operation': operation,
                'customers': customers,
                'sales_prices': sales_prices,
                'days': days,
                'items': items,
                'sub_items': sub_items
            })
    return render(request, 'operation/forms/form.html', {
        'form': form,
        'operation': operation,
        'page_title': 'Tur Düzenle',
        'post_url': reverse('tour:sub_item_tour_update', args=[sub_item.id])
    })

def sub_item_transfer_create(request, operation_item_id):
    item = get_object_or_404(OperationItem, id=operation_item_id)
    operation = get_object_or_404(Operation, id=item.operation_day.operation.id)
    customers = OperationCustomer.objects.filter(operation=operation).select_related('operation')
    sales_prices = OperationSalesPrice.objects.filter(operation=operation).select_related('operation')
    days = OperationDay.objects.filter(operation=operation).select_related('operation')
    items = OperationItem.objects.filter(operation_day__in=days).select_related(
        'operation_day',
        'vehicle_type',
        'vehicle_supplier',
        'no_vehicle_tour',
        'no_vehicle_activity',
        'activity_supplier'
    )
    sub_items = OperationSubItem.objects.filter(operation_item__in=items).select_related(
        'operation_item',
        'tour',
        'transfer',
        'hotel',
        'guide',
        'activity',
        'activity_supplier'
    )
    form = OperationSubItemTransferForm()
    if request.method == 'POST':
        form = OperationSubItemTransferForm(request.POST)
        if form.is_valid():
            transfer = form.save(commit=False)
            transfer.operation_item = item
            transfer.subitem_type = "TRANSFER"
            transfer.save()
            return render(request, 'operation/includes/operation_detail.html', {
                'form': form,
                'operation': operation,
                'customers': customers,
                'sales_prices': sales_prices,
                'days': days,
                'items': items,
                'sub_items': sub_items
            })
    return render(request, 'operation/forms/form.html', {
        'form': form,
        'operation': operation,
        'page_title': 'Transfer Ekle',
        'post_url': reverse('tour:sub_item_transfer_create', args=[item.id])
    })

def sub_item_transfer_update(request, operation_sub_item_id):
    sub_item = get_object_or_404(OperationSubItem, id=operation_sub_item_id)
    operation = get_object_or_404(Operation, id=sub_item.operation_item.operation_day.operation.id)
    customers = OperationCustomer.objects.filter(operation=operation).select_related('operation')
    sales_prices = OperationSalesPrice.objects.filter(operation=operation).select_related('operation')
    days = OperationDay.objects.filter(operation=operation).select_related('operation')
    items = OperationItem.objects.filter(operation_day__in=days).select_related(
        'operation_day',
        'vehicle_type',
        'vehicle_supplier',
        'no_vehicle_tour',
        'no_vehicle_activity',
        'activity_supplier'
    )
    sub_items = OperationSubItem.objects.filter(operation_item__in=items).select_related(
        'operation_item',
        'tour',
        'transfer',
        'hotel',
        'guide',
        'activity',
        'activity_supplier'
    )
    form = OperationSubItemTransferForm(instance=sub_item)
    if request.method == 'POST':
        form = OperationSubItemTransferForm(request.POST, instance=sub_item)
        if form.is_valid():
            form.save()
            return render(request, 'operation/includes/operation_detail.html', {
                'form': form,
                'operation': operation,
                'customers': customers,
                'sales_prices': sales_prices,
                'days': days,
                'items': items,
                'sub_items': sub_items
            })
    return render(request, 'operation/forms/form.html', {
        'form': form,
        'operation': operation,
        'page_title': 'Transfer Düzenle',
        'post_url': reverse('tour:sub_item_transfer_update', args=[sub_item.id])
    })

def sub_item_hotel_create(request, operation_item_id):
    item = get_object_or_404(OperationItem, id=operation_item_id)
    operation = get_object_or_404(Operation, id=item.operation_day.operation.id)
    customers = OperationCustomer.objects.filter(operation=operation).select_related('operation')
    sales_prices = OperationSalesPrice.objects.filter(operation=operation).select_related('operation')
    days = OperationDay.objects.filter(operation=operation).select_related('operation')
    items = OperationItem.objects.filter(operation_day__in=days).select_related(
        'operation_day',
        'vehicle_type',
        'vehicle_supplier',
        'no_vehicle_tour',
        'no_vehicle_activity',
        'activity_supplier'
    )
    sub_items = OperationSubItem.objects.filter(operation_item__in=items).select_related(
        'operation_item',
        'tour',
        'transfer',
        'hotel',
        'guide',
        'activity',
        'activity_supplier'
    )
    form = OperationSubItemHotelForm()
    if request.method == 'POST':
        form = OperationSubItemHotelForm(request.POST)
        if form.is_valid():
            hotel = form.save(commit=False)
            hotel.operation_item = item
            hotel.subitem_type = "HOTEL"
            hotel.save()
            return render(request, 'operation/includes/operation_detail.html', {
                'form': form,
                'operation': operation,
                'customers': customers,
                'sales_prices': sales_prices,
                'days': days,
                'items': items,
                'sub_items': sub_items
            })
    return render(request, 'operation/forms/form.html', {
        'form': form,
        'operation': operation,
        'page_title': 'Otel Ekle',
        'post_url': reverse('tour:sub_item_hotel_create', args=[item.id])
    })

def sub_item_hotel_update(request, operation_sub_item_id):
    sub_item = get_object_or_404(OperationSubItem, id=operation_sub_item_id)
    operation = get_object_or_404(Operation, id=sub_item.operation_item.operation_day.operation.id)
    customers = OperationCustomer.objects.filter(operation=operation).select_related('operation')
    sales_prices = OperationSalesPrice.objects.filter(operation=operation).select_related('operation')
    days = OperationDay.objects.filter(operation=operation).select_related('operation')
    items = OperationItem.objects.filter(operation_day__in=days).select_related(
        'operation_day',
        'vehicle_type',
        'vehicle_supplier',
        'no_vehicle_tour',
        'no_vehicle_activity',
        'activity_supplier'
    )
    sub_items = OperationSubItem.objects.filter(operation_item__in=items).select_related(
        'operation_item',
        'tour',
        'transfer',
        'hotel',
        'guide',
        'activity',
        'activity_supplier'
    )
    form = OperationSubItemHotelForm(instance=sub_item)
    if request.method == 'POST':
        form = OperationSubItemHotelForm(request.POST, instance=sub_item)
        if form.is_valid():
            form.save()
            return render(request, 'operation/includes/operation_detail.html', {
                'form': form,
                'operation': operation,
                'customers': customers,
                'sales_prices': sales_prices,
                'days': days,
                'items': items,
                'sub_items': sub_items
            })
    return render(request, 'operation/forms/form.html', {
        'form': form,
        'operation': operation,
        'page_title': 'Otel Düzenle',
        'post_url': reverse('tour:sub_item_hotel_update', args=[sub_item.id])
    })

def sub_item_activity_create(request, operation_item_id):
    item = get_object_or_404(OperationItem, id=operation_item_id)
    operation = get_object_or_404(Operation, id=item.operation_day.operation.id)
    customers = OperationCustomer.objects.filter(operation=operation).select_related('operation')
    sales_prices = OperationSalesPrice.objects.filter(operation=operation).select_related('operation')
    days = OperationDay.objects.filter(operation=operation).select_related('operation')
    items = OperationItem.objects.filter(operation_day__in=days).select_related(
        'operation_day',
        'vehicle_type',
        'vehicle_supplier',
        'no_vehicle_tour',
        'no_vehicle_activity',
        'activity_supplier'
    )
    sub_items = OperationSubItem.objects.filter(operation_item__in=items).select_related(
        'operation_item',
        'tour',
        'transfer',
        'hotel',
        'guide',
        'activity',
        'activity_supplier'
    )
    form = OperationSubItemActivityForm()
    if request.method == 'POST':
        form = OperationSubItemActivityForm(request.POST)
        if form.is_valid():
            activity = form.save(commit=False)
            activity.operation_item = item
            activity.subitem_type = "ACTIVITY"
            activity.save()
            return render(request, 'operation/includes/operation_detail.html', {
                'form': form,
                'operation': operation,
                'customers': customers,
                'sales_prices': sales_prices,
                'days': days,
                'items': items,
                'sub_items': sub_items
            })
    return render(request, 'operation/forms/form.html', {
        'form': form,
        'operation': operation,
        'page_title': 'Aktivite Ekle',
        'post_url': reverse('tour:sub_item_activity_create', args=[item.id])
    })

def sub_item_activity_update(request, operation_sub_item_id):
    sub_item = get_object_or_404(OperationSubItem, id=operation_sub_item_id)
    operation = get_object_or_404(Operation, id=sub_item.operation_item.operation_day.operation.id)
    customers = OperationCustomer.objects.filter(operation=operation).select_related('operation')
    sales_prices = OperationSalesPrice.objects.filter(operation=operation).select_related('operation')
    days = OperationDay.objects.filter(operation=operation).select_related('operation')
    items = OperationItem.objects.filter(operation_day__in=days).select_related(
        'operation_day',
        'vehicle_type',
        'vehicle_supplier',
        'no_vehicle_tour',
        'no_vehicle_activity',
        'activity_supplier'
    )
    sub_items = OperationSubItem.objects.filter(operation_item__in=items).select_related(
        'operation_item',
        'tour',
        'transfer',
        'hotel',
        'guide',
        'activity',
        'activity_supplier'
    )
    form = OperationSubItemActivityForm(instance=sub_item)
    if request.method == 'POST':
        form = OperationSubItemActivityForm(request.POST, instance=sub_item)
        if form.is_valid():
            form.save()
            return render(request, 'operation/includes/operation_detail.html', {
                'form': form,
                'operation': operation,
                'customers': customers,
                'sales_prices': sales_prices,
                'days': days,
                'items': items,
                'sub_items': sub_items
            })
    return render(request, 'operation/forms/form.html', {
        'form': form,
        'operation': operation,
        'page_title': 'Aktivite Düzenle',
        'post_url': reverse('tour:sub_item_activity_update', args=[sub_item.id])
    })

def sub_item_museum_create(request, operation_item_id):
    item = get_object_or_404(OperationItem, id=operation_item_id)
    operation = get_object_or_404(Operation, id=item.operation_day.operation.id)
    customers = OperationCustomer.objects.filter(operation=operation).select_related('operation')
    sales_prices = OperationSalesPrice.objects.filter(operation=operation).select_related('operation')
    days = OperationDay.objects.filter(operation=operation).select_related('operation')
    items = OperationItem.objects.filter(operation_day__in=days).select_related(
        'operation_day',
        'vehicle_type',
        'vehicle_supplier',
        'no_vehicle_tour',
        'no_vehicle_activity',
        'activity_supplier'
    )
    sub_items = OperationSubItem.objects.filter(operation_item__in=items).select_related(
        'operation_item',
        'tour',
        'transfer',
        'hotel',
        'guide',
        'activity',
        'activity_supplier'
    ).prefetch_related('museums')
    form = OperationSubItemMuseumForm()
    if request.method == 'POST':
        form = OperationSubItemMuseumForm(request.POST)
        if form.is_valid():
            museum = form.save(commit=False)
            museum.operation_item = item
            museum.subitem_type = "MUSEUM"
            museum.save()
            museum.museums.set(form.cleaned_data['museums'])
            return render(request, 'operation/includes/operation_detail.html', {
                'form': form,
                'operation': operation,
                'customers': customers,
                'sales_prices': sales_prices,
                'days': days,
                'items': items,
                'sub_items': sub_items
            })
        else:
            print(form.errors)

    return render(request, 'operation/forms/form.html', {
        'form': form,
        'operation': operation,
        'page_title': 'Müze Ekle',
        'post_url': reverse('tour:sub_item_museum_create', args=[item.id])
    })

def sub_item_museum_update(request, operation_sub_item_id):
    sub_item = get_object_or_404(OperationSubItem, id=operation_sub_item_id)
    operation = get_object_or_404(Operation, id=sub_item.operation_item.operation_day.operation.id)
    customers = OperationCustomer.objects.filter(operation=operation).select_related('operation')
    sales_prices = OperationSalesPrice.objects.filter(operation=operation).select_related('operation')
    days = OperationDay.objects.filter(operation=operation).select_related('operation')
    items = OperationItem.objects.filter(operation_day__in=days).select_related(
        'operation_day',
        'vehicle_type',
        'vehicle_supplier',
        'no_vehicle_tour',
        'no_vehicle_activity',
        'activity_supplier'
    )
    sub_items = OperationSubItem.objects.filter(operation_item__in=items).select_related(
        'operation_item',
        'tour',
        'transfer',
        'hotel',
        'guide',
        'activity',
        'activity_supplier'
    ).prefetch_related('museums')
    form = OperationSubItemMuseumForm(instance=sub_item)
    if request.method == 'POST':
        form = OperationSubItemMuseumForm(request.POST, instance=sub_item)
        if form.is_valid():
            museum = form.save()
            museum.museums.set(form.cleaned_data['museums'])
            return render(request, 'operation/includes/operation_detail.html', {
                'form': form,
                'operation': operation,
                'customers': customers,
                'sales_prices': sales_prices,
                'days': days,
                'items': items,
                'sub_items': sub_items
            })
        else:
            print(form.errors)
    return render(request, 'operation/forms/form.html', {
        'form': form,
        'operation': operation,
        'page_title': 'Müze Düzenle',
        'post_url': reverse('tour:sub_item_museum_update', args=[sub_item.id])
    })

def sub_item_guide_create(request, operation_item_id):
    item = get_object_or_404(OperationItem, id=operation_item_id)
    operation = get_object_or_404(Operation, id=item.operation_day.operation.id)
    customers = OperationCustomer.objects.filter(operation=operation).select_related('operation')
    sales_prices = OperationSalesPrice.objects.filter(operation=operation).select_related('operation')
    days = OperationDay.objects.filter(operation=operation).select_related('operation')
    items = OperationItem.objects.filter(operation_day__in=days).select_related(
        'operation_day',
        'vehicle_type',
        'vehicle_supplier',
        'no_vehicle_tour',
        'no_vehicle_activity',
        'activity_supplier'
    )
    sub_items = OperationSubItem.objects.filter(operation_item__in=items).select_related(
        'operation_item',
        'tour',
        'transfer',
        'hotel',
        'guide',
        'activity',
        'activity_supplier'
    )
    form = OperationSubItemGuideForm()
    if request.method == 'POST':
        form = OperationSubItemGuideForm(request.POST)
        if form.is_valid():
            guide = form.save(commit=False)
            guide.operation_item = item
            guide.subitem_type = "GUIDE"
            guide.save()
            return render(request, 'operation/includes/operation_detail.html', {
                'form': form,
                'operation': operation,
                'customers': customers,
                'sales_prices': sales_prices,
                'days': days,
                'items': items,
                'sub_items': sub_items
            })
    return render(request, 'operation/forms/form.html', {
        'form': form,
        'operation': operation,
        'page_title': 'Rehber Ekle',
        'post_url': reverse('tour:sub_item_guide_create', args=[item.id])
    })

def sub_item_guide_update(request, operation_sub_item_id):
    sub_item = get_object_or_404(OperationSubItem, id=operation_sub_item_id)
    operation = get_object_or_404(Operation, id=sub_item.operation_item.operation_day.operation.id)
    customers = OperationCustomer.objects.filter(operation=operation).select_related('operation')
    sales_prices = OperationSalesPrice.objects.filter(operation=operation).select_related('operation')
    days = OperationDay.objects.filter(operation=operation).select_related('operation')
    items = OperationItem.objects.filter(operation_day__in=days).select_related(
        'operation_day',
        'vehicle_type',
        'vehicle_supplier',
        'no_vehicle_tour',
        'no_vehicle_activity',
        'activity_supplier'
    )
    sub_items = OperationSubItem.objects.filter(operation_item__in=items).select_related(
        'operation_item',
        'tour',
        'transfer',
        'hotel',
        'guide',
        'activity',
        'activity_supplier'
    )
    form = OperationSubItemGuideForm(instance=sub_item)
    if request.method == 'POST':
        form = OperationSubItemGuideForm(request.POST, instance=sub_item)
        if form.is_valid():
            form.save()
            return render(request, 'operation/includes/operation_detail.html', {
                'form': form,
                'operation': operation,
                'customers': customers,
                'sales_prices': sales_prices,
                'days': days,
                'items': items,
                'sub_items': sub_items
            })
    return render(request, 'operation/forms/form.html', {
        'form': form,
        'operation': operation,
        'page_title': 'Rehber Düzenle',
        'post_url': reverse('tour:sub_item_guide_update', args=[sub_item.id])
    })

def sub_item_other_price_create(request, operation_item_id):
    item = get_object_or_404(OperationItem, id=operation_item_id)
    operation = get_object_or_404(Operation, id=item.operation_day.operation.id)
    customers = OperationCustomer.objects.filter(operation=operation).select_related('operation')
    sales_prices = OperationSalesPrice.objects.filter(operation=operation).select_related('operation')
    days = OperationDay.objects.filter(operation=operation).select_related('operation')
    items = OperationItem.objects.filter(operation_day__in=days).select_related(
        'operation_day',
        'vehicle_type',
        'vehicle_supplier',
        'no_vehicle_tour',
        'no_vehicle_activity',
        'activity_supplier'
    )
    sub_items = OperationSubItem.objects.filter(operation_item__in=items).select_related(
        'operation_item',
        'tour',
        'transfer',
        'hotel',
        'guide',
        'activity',
        'activity_supplier'
    )
    form = OperationSubItemOtherPriceForm()
    if request.method == 'POST':
        form = OperationSubItemOtherPriceForm(request.POST)
        if form.is_valid():
            other_price = form.save(commit=False)
            other_price.operation_item = item
            other_price.subitem_type = "OTHER_PRICE"
            other_price.save()
            return render(request, 'operation/includes/operation_detail.html', {
                'form': form,
                'operation': operation,
                'customers': customers,
                'sales_prices': sales_prices,
                'days': days,
                'items': items,
                'sub_items': sub_items
            })
    return render(request, 'operation/forms/form.html', {
        'form': form,
        'operation': operation,
        'page_title': 'Masraf Ekle',
        'post_url': reverse('tour:sub_item_other_price_create', args=[item.id])
    })

def sub_item_other_price_update(request, operation_sub_item_id):
    sub_item = get_object_or_404(OperationSubItem, id=operation_sub_item_id)
    operation = get_object_or_404(Operation, id=sub_item.operation_item.operation_day.operation.id)
    customers = OperationCustomer.objects.filter(operation=operation).select_related('operation')
    sales_prices = OperationSalesPrice.objects.filter(operation=operation).select_related('operation')
    days = OperationDay.objects.filter(operation=operation).select_related('operation')
    items = OperationItem.objects.filter(operation_day__in=days).select_related(
        'operation_day',
        'vehicle_type',
        'vehicle_supplier',
        'no_vehicle_tour',
        'no_vehicle_activity',
        'activity_supplier'
    )
    sub_items = OperationSubItem.objects.filter(operation_item__in=items).select_related(
        'operation_item',
        'tour',
        'transfer',
        'hotel',
        'guide',
        'activity',
        'activity_supplier'
    )
    form = OperationSubItemOtherPriceForm(instance=sub_item)
    if request.method == 'POST':
        form = OperationSubItemOtherPriceForm(request.POST, instance=sub_item)
        if form.is_valid():
            form.save()
            return render(request, 'operation/includes/operation_detail.html', {
                'form': form,
                'operation': operation,
                'customers': customers,
                'sales_prices': sales_prices,
                'days': days,
                'items': items,
                'sub_items': sub_items
            })
    return render(request, 'operation/forms/form.html', {
        'form': form,
        'operation': operation,
        'page_title': 'Masraf Düzenle',
        'post_url': reverse('tour:sub_item_other_price_update', args=[sub_item.id])
    })


def operation_customer_create(request, operation_id):
    operation = get_object_or_404(Operation, id=operation_id)
    customers = OperationCustomer.objects.filter(operation=operation).select_related('operation')
    sales_prices = OperationSalesPrice.objects.filter(operation=operation).select_related('operation')
    days = OperationDay.objects.filter(operation=operation).select_related('operation')
    items = OperationItem.objects.filter(operation_day__in=days).select_related(
        'operation_day',
        'vehicle_type',
        'vehicle_supplier',
        'no_vehicle_tour',
        'no_vehicle_activity',
        'activity_supplier'
    )
    sub_items = OperationSubItem.objects.filter(operation_item__in=items).select_related(
        'operation_item',
        'tour',
        'transfer',
        'hotel',
        'guide',
        'activity',
        'activity_supplier'
    ).prefetch_related('museums')
    form = OperationCustomerForm()
    if request.method == 'POST':
        form = OperationCustomerForm(request.POST)
        if form.is_valid():
            customer = form.save(commit=False)
            customer.operation = operation
            customer.save()
            return render(request, 'operation/includes/operation_detail.html', {
                'form': form,
                'operation': operation,
                'customers': customers,
                'sales_prices': sales_prices,
                'days': days,
                'items': items,
                'sub_items': sub_items
            })
    return render(request, 'operation/forms/form.html', {
        'form': form,
        'operation': operation,
        'page_title': 'Müşteri Ekle',
        'post_url': reverse('tour:operation_customer_create', args=[operation.id])
    })


def operation_sales_price_create(request, operation_id):
    operation = get_object_or_404(Operation, id=operation_id)
    customers = OperationCustomer.objects.filter(operation=operation).select_related('operation')
    sales_prices = OperationSalesPrice.objects.filter(operation=operation).select_related('operation')
    days = OperationDay.objects.filter(operation=operation).select_related('operation')
    items = OperationItem.objects.filter(operation_day__in=days).select_related(
        'operation_day',
        'vehicle_type',
        'vehicle_supplier',
        'no_vehicle_tour',
        'no_vehicle_activity',
        'activity_supplier'
    )
    sub_items = OperationSubItem.objects.filter(operation_item__in=items).select_related(
        'operation_item',
        'tour',
        'transfer',
        'hotel',
        'guide',
        'activity',
        'activity_supplier'
    )
    form = OperationSalesPriceForm()
    if request.method == 'POST':
        form = OperationSalesPriceForm(request.POST)
        if form.is_valid():
            sales_price = form.save(commit=False)
            sales_price.operation = operation
            sales_price.save()
            return render(request, 'operation/includes/operation_detail.html', {
                'form': form,
                'operation': operation,
                'customers': customers,
                'sales_prices': sales_prices,
                'days': days,
                'items': items,
                'sub_items': sub_items
            })
    return render(request, 'operation/forms/form.html', {
        'form': form,
        'operation': operation,
        'page_title': 'Satış Fiyatı Ekle',
        'post_url': reverse('tour:operation_sales_price_create', args=[operation.id])
    })

def operation_create(request):
    form = OperationForm()
    if request.method == 'POST':
        form = OperationForm(request.POST)
        if form.is_valid():
            operation = form.save(commit=False)
            operation.created_by = request.user
            operation.save()
            return redirect('tour:operation', operation_id=operation.id)
    return render(request, 'operation/operation_create.html', {
        'form': form,
        'page_title': 'Operasyon Ekle',
        'post_url': reverse('tour:operation_create')
    })


def operation_list(request):
    # Filtreleme parametrelerini al
    reference_number = request.GET.get('reference_number', '')
    created_by = request.GET.get('created_by', '')
    follow_by = request.GET.get('follow_by', '')
    buyer_company = request.GET.get('buyer_company', '')
    status = request.GET.get('status', '')
    month = request.GET.get('month', '')

    # Temel sorguyu oluştur
    operations = Operation.objects.select_related(
        'created_by',
        'follow_by',
        'buyer_company'
    )

    # Ay filtresi
    if month:
        operations = operations.filter(start_date__month=month)
    else:
        # Ay seçilmemişse mevcut ayı göster
        current_month = datetime.now().month
        operations = operations.filter(start_date__month=current_month)

    # Diğer filtreleri uygula
    if reference_number:
        operations = operations.filter(reference_number__icontains=reference_number)
    if created_by:
        operations = operations.filter(created_by__username=created_by)
    if follow_by:
        operations = operations.filter(follow_by__username=follow_by)
    if buyer_company:
        operations = operations.filter(buyer_company__name=buyer_company)
    if status:
        operations = operations.filter(status=status)

    # Filtreleme için gerekli verileri al
    buyer_companies = BuyerCompany.objects.all()
    users = CustomUser.objects.all()
    status_choices = Operation.STATUS_CHOICES

    # Ay seçenekleri
    months = [
        (1, 'Ocak'), (2, 'Şubat'), (3, 'Mart'), (4, 'Nisan'),
        (5, 'Mayıs'), (6, 'Haziran'), (7, 'Temmuz'), (8, 'Ağustos'),
        (9, 'Eylül'), (10, 'Ekim'), (11, 'Kasım'), (12, 'Aralık')
    ]

    context = {
        'operations': operations,
        'buyer_companies': buyer_companies,
        'users': users,
        'status_choices': status_choices,
        'months': months,
        'current_month': int(month) if month else datetime.now().month,
        'filters': {
            'reference_number': reference_number,
            'created_by': created_by,
            'follow_by': follow_by,
            'buyer_company': buyer_company,
            'status': status,
            'month': month
        }
    }

    return render(request, 'operation/operation_list.html', context)


def operation_jobs(request):
    # Bugünün tarihini al
    today = datetime.now().date()
    
    # Bugünden başlayarak 7 günlük tarih aralığını oluştur
    date_range = [today + timedelta(days=i) for i in range(7)]
    
    # URL'den seçilen tarihi al, yoksa bugünü kullan
    selected_date_str = request.GET.get('date')
    if selected_date_str:
        try:
            selected_date = datetime.strptime(selected_date_str, '%Y-%m-%d').date()
            # Seçilen tarih 7 günlük aralıkta değilse bugünü kullan
            if selected_date not in date_range:
                selected_date = today
        except ValueError:
            selected_date = today
    else:
        selected_date = today
    
    # Her gün için operasyon günlerini ve ilişkili verileri al
    days = OperationDay.objects.filter(
        date__in=date_range
    ).select_related(
        'operation',
        'operation__created_by',
        'operation__follow_by',
        'operation__buyer_company'
    ).prefetch_related(
        Prefetch(
            'items',
            queryset=OperationItem.objects.select_related(
                'vehicle_type',
                'vehicle_supplier',
                'no_vehicle_tour',
                'no_vehicle_activity',
                'activity_supplier'
            ).prefetch_related(
                Prefetch(
                    'subitems',
                    queryset=OperationSubItem.objects.select_related(
                        'tour',
                        'transfer',
                        'hotel',
                        'guide',
                        'activity',
                        'activity_supplier'
                    ).prefetch_related('museums')
                )
            )
        )
    )
    
    context = {
        'date_range': date_range,
        'days': days,
        'today': today,
        'selected_date': selected_date
    }
    
    return render(request, 'operation/operation_jobs.html', context)

def jobs_vehicle_item_update(request, operation_item_id):
    item = get_object_or_404(OperationItem, id=operation_item_id)
    form = OperationItemVehicleForm(instance=item)
    if request.method == 'POST':
        form = OperationItemVehicleForm(request.POST, instance=item)
        if form.is_valid():
            form.save()
            return render(request, 'operation/includes/jobs_item.html', {
                'form': form,
                'item': item
            })
    return render(request, 'operation/forms/jobs_item_form.html', {
        'form': form,
        'item': item,
        'page_title': 'Araç Düzenle',
        'post_url': reverse('tour:jobs_vehicle_item_update', args=[item.id])
    })

def jobs_no_vehicle_tour_item_update(request, operation_item_id):
    item = get_object_or_404(OperationItem, id=operation_item_id)
    form = OperationItemNoVehicleTourForm(instance=item)
    if request.method == 'POST':
        form = OperationItemNoVehicleTourForm(request.POST, instance=item)
        if form.is_valid():
            form.save()
            return render(request, 'operation/includes/jobs_item.html', {
                'form': form,
                'item': item
            })
    return render(request, 'operation/forms/jobs_item_form.html', {
        'form': form,
        'item': item,
        'page_title': 'Araçsız Tur Düzenle',
        'post_url': reverse('tour:jobs_no_vehicle_tour_item_update', args=[item.id])
    })

def jobs_no_vehicle_activity_item_update(request, operation_item_id):
    item = get_object_or_404(OperationItem, id=operation_item_id)
    form = OperationItemActivityForm(instance=item)
    if request.method == 'POST':
        form = OperationItemActivityForm(request.POST, instance=item)
        if form.is_valid():
            form.save()
            return render(request, 'operation/includes/jobs_item.html', {
                'form': form,
                'item': item
            })
    return render(request, 'operation/forms/jobs_item_form.html', {
        'form': form,
        'item': item,
        'page_title': 'Araçsız Aktivite Düzenle',
        'post_url': reverse('tour:jobs_no_vehicle_activity_item_update', args=[item.id])
    })

def jobs_no_vehicle_guide_item_update(request, operation_item_id):
    item = get_object_or_404(OperationItem, id=operation_item_id)
    form = OperationItemNoVehicleGuideForm(instance=item)
    if request.method == 'POST':
        form = OperationItemNoVehicleGuideForm(request.POST, instance=item)
        if form.is_valid():
            form.save()
            return render(request, 'operation/includes/jobs_item.html', {
                'form': form,
                'item': item
            })
    return render(request, 'operation/forms/jobs_item_form.html', {
        'form': form,
        'item': item,
        'page_title': 'Araçsız Rehber Düzenle',
        'post_url': reverse('tour:jobs_no_vehicle_guide_item_update', args=[item.id])
    })


def jobs_sub_item_tour_update(request, operation_sub_item_id):
    sub_item = get_object_or_404(OperationSubItem, id=operation_sub_item_id)
    form = OperationSubItemTourForm(instance=sub_item)
    if request.method == 'POST':
        form = OperationSubItemTourForm(request.POST, instance=sub_item)
        if form.is_valid():
            form.save()
            return render(request, 'operation/includes/jobs_subitem.html', {
                'form': form,
                'subitem': sub_item
            })
    return render(request, 'operation/forms/jobs_subitem_form.html', {
        'form': form,
        'subitem': sub_item,
        'page_title': 'Tur Düzenle',
        'post_url': reverse('tour:jobs_sub_item_tour_update', args=[sub_item.id])
    })

def jobs_sub_item_transfer_update(request, operation_sub_item_id):
    sub_item = get_object_or_404(OperationSubItem, id=operation_sub_item_id)
    form = OperationSubItemTransferForm(instance=sub_item)
    if request.method == 'POST':
        form = OperationSubItemTransferForm(request.POST, instance=sub_item)
        if form.is_valid():
            form.save()
            return render(request, 'operation/includes/jobs_subitem.html', {
                'form': form,
                'subitem': sub_item
            })
    return render(request, 'operation/forms/jobs_subitem_form.html', {
        'form': form,
        'subitem': sub_item,
        'page_title': 'Transfer Düzenle',
        'post_url': reverse('tour:jobs_sub_item_transfer_update', args=[sub_item.id])
    })


def jobs_sub_item_hotel_update(request, operation_sub_item_id):
    sub_item = get_object_or_404(OperationSubItem, id=operation_sub_item_id)
    form = OperationSubItemHotelForm(instance=sub_item)
    if request.method == 'POST':
        form = OperationSubItemHotelForm(request.POST, instance=sub_item)
        if form.is_valid():
            form.save()
            return render(request, 'operation/includes/jobs_subitem.html', {
                'form': form,
                'subitem': sub_item
            })
    return render(request, 'operation/forms/jobs_subitem_form.html', {
        'form': form,
        'subitem': sub_item,
        'page_title': 'Otel Düzenle',
        'post_url': reverse('tour:jobs_sub_item_hotel_update', args=[sub_item.id])
    })  



def jobs_sub_item_activity_update(request, operation_sub_item_id):
    sub_item = get_object_or_404(OperationSubItem, id=operation_sub_item_id)
    form = OperationSubItemActivityForm(instance=sub_item)
    if request.method == 'POST':
        form = OperationSubItemActivityForm(request.POST, instance=sub_item)
        if form.is_valid():
            form.save()
            return render(request, 'operation/includes/jobs_subitem.html', {
                'form': form,
                'subitem': sub_item
            })
    return render(request, 'operation/forms/jobs_subitem_form.html', {
        'form': form,
        'subitem': sub_item,
        'page_title': 'Aktivite Düzenle',
        'post_url': reverse('tour:jobs_sub_item_activity_update', args=[sub_item.id])
    })

def jobs_sub_item_museum_update(request, operation_sub_item_id):
    sub_item = get_object_or_404(OperationSubItem, id=operation_sub_item_id)
    form = OperationSubItemMuseumForm(instance=sub_item)
    if request.method == 'POST':
        form = OperationSubItemMuseumForm(request.POST, instance=sub_item)
        if form.is_valid():
            form.save()
            return render(request, 'operation/includes/jobs_subitem.html', {
                'form': form,
                'subitem': sub_item
            })
    return render(request, 'operation/forms/jobs_subitem_form.html', {
        'form': form,
        'subitem': sub_item,
        'page_title': 'Müzeler Düzenle',
        'post_url': reverse('tour:jobs_sub_item_museum_update', args=[sub_item.id])
    })

def jobs_sub_item_guide_update(request, operation_sub_item_id):
    sub_item = get_object_or_404(OperationSubItem, id=operation_sub_item_id)
    form = OperationSubItemGuideForm(instance=sub_item)
    if request.method == 'POST':
        form = OperationSubItemGuideForm(request.POST, instance=sub_item)
        if form.is_valid():
            form.save()
            return render(request, 'operation/includes/jobs_subitem.html', {
                'form': form,
                'subitem': sub_item
            })
    return render(request, 'operation/forms/jobs_subitem_form.html', {
        'form': form,
        'subitem': sub_item,
        'page_title': 'Rehber Düzenle',
        'post_url': reverse('tour:jobs_sub_item_guide_update', args=[sub_item.id])
    })



def jobs_sub_item_other_price_update(request, operation_sub_item_id):
    sub_item = get_object_or_404(OperationSubItem, id=operation_sub_item_id)
    form = OperationSubItemOtherPriceForm(instance=sub_item)
    if request.method == 'POST':
        form = OperationSubItemOtherPriceForm(request.POST, instance=sub_item)
        if form.is_valid():
            form.save()
            return render(request, 'operation/includes/jobs_subitem.html', {
                'form': form,
                'subitem': sub_item
            })
    return render(request, 'operation/forms/jobs_subitem_form.html', {
        'form': form,
        'subitem': sub_item,
        'page_title': 'Ekstra Masraf Düzenle',
        'post_url': reverse('tour:jobs_sub_item_other_price_update', args=[sub_item.id])
    })



def my_operation_jobs(request):
    # Bugünün tarihini al
    today = datetime.now().date()
    
    # Bugünden başlayarak 7 günlük tarih aralığını oluştur
    date_range = [today + timedelta(days=i) for i in range(7)]
    
    # URL'den seçilen tarihi al, yoksa bugünü kullan
    selected_date_str = request.GET.get('date')
    if selected_date_str:
        try:
            selected_date = datetime.strptime(selected_date_str, '%Y-%m-%d').date()
            # Seçilen tarih 7 günlük aralıkta değilse bugünü kullan
            if selected_date not in date_range:
                selected_date = today
        except ValueError:
            selected_date = today
    else:
        selected_date = today
    
    # Her gün için operasyon günlerini ve ilişkili verileri al
    days = OperationDay.objects.filter(
        date__in=date_range, 
        operation__follow_by=request.user
    ).select_related(
        'operation',
        'operation__created_by',
        'operation__follow_by',
        'operation__buyer_company'
    ).prefetch_related(
        Prefetch(
            'items',
            queryset=OperationItem.objects.select_related(
                'vehicle_type',
                'vehicle_supplier',
                'no_vehicle_tour',
                'no_vehicle_activity',
                'activity_supplier'
            ).prefetch_related(
                Prefetch(
                    'subitems',
                    queryset=OperationSubItem.objects.select_related(
                        'tour',
                        'transfer',
                        'hotel',
                        'guide',
                        'activity',
                        'activity_supplier'
                    ).prefetch_related('museums')
                )
            )
        )
    )
    
    context = {
        'date_range': date_range,
        'days': days,
        'today': today,
        'selected_date': selected_date
    }
    
    return render(request, 'operation/operation_jobs.html', context)


